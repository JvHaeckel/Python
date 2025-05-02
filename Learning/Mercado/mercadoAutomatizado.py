# Exercício da página: 
# https://community.revelo.com.br/como-utilizar-python-para-automatizar-planilhas-do-excel/

# Automatizando o relatório em Excel utilizando Python Agora que temos um relatório apresentado por meio de uma
# Tabela Dinâmica, a próxima e mais importante parte é automatização da sua criação. Então, na próxima vez que 
# você quiser fazer esse relatório, basta digitar o nome do arquivo e executá-lo com o código Python.

import pandas as pd
from openpyxl import load_workbook

def automate_excel():
    # Caminho completo do arquivo Excel
    file_name = r'C:\Users\joaorocha\Desktop\Py\Learning\Mercado\supermarket_sales.xlsx'

    # Ler o arquivo Excel
    excel_file = pd.read_excel(file_name)

    # Criar a Tabela Dinâmica (Pivot Table)
    report_table = excel_file.pivot_table(
        index='Gender',
        columns='Product line',
        values='Total',
        aggfunc='sum'
    ).round(0)

    # Extrai a parte "sales.xlsx", ou seja, o nome do mês (sales) e a extensão do arquivo.
    month_and_extension = file_name.split('_')[1]

    # Exportar a tabela para um novo arquivo Excel
    report_table.to_excel(
        f'report_{month_and_extension}', # Você está dizendo ao Python: Salve o arquivo com o nome "report_" + o conteúdo da 
        #variável month_and_extension. Então o nome do arquivo final será: report_sales.xlsx
        sheet_name='Report',  # nomeia a planilha
        startrow = 4 #  Começa a escrever a tabela na linha 5 do Excel (porque o índice começa em zero).Isso deixa um espaço de
                     # 4 linhas em branco no topo, que pode ser usado depois para título, instruções ou cabeçalho manual, por exemplo
    )

    # Abrir o novo arquivo para edição com openpyxl
    wb = load_workbook(f'report_{month_and_extension}')
    sheet = wb['Report']

    # Pegar as dimensões da planilha ativa
    min_column = sheet.min_column
    max_column = sheet.max_column
    min_row = sheet.min_row
    max_row = sheet.max_row

    # Extrair o nome do mês (sem a extensão)
    month_name = month_and_extension.split('.')[0]

    # Salvar as alterações no arquivo Excel
    wb.save(f'report_{month_and_extension}')

# Executar a função
automate_excel()