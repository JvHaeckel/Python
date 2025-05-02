# Exercício da página: 
# https://community.revelo.com.br/como-utilizar-python-para-automatizar-planilhas-do-excel/

# Toda vez que colocar os imports está pedindo para baixar então coloque no terminal python: pip install openpyxl

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

# Para ler nosso arquivo ““supermarket_sales.xlsx”, usaremos a função pd.read_excel(), criando uma variável e colando 
#  caminho da pasta, lembrando que precisa colocar o r do lado de fora pq é raw arquivo.

table = pd.read_excel(r"C:\Users\joaorocha\Desktop\Py\Learning\Mercado\supermarket_sales.xlsx")
# print(table[['Gender', 'Product line', 'Total']])

#Para criar nossa Tabela Dinâmica, usaremos a função .pivot_table() para mostrar, por exemplo, o dinheiro total 
# gasto por homens e mulheres nas diferentes linhas de produtos.

total_gasto = table.pivot_table (
                                 index='Gender', 
                                 columns='Product line', 
                                 values='Total', 
                                 aggfunc= sum).round(0)

# ✅ index='Gender' Isso define que as linhas da tabela serão separadas pelo gênero (Male ou Female).

# ✅ columns='Product line' as colunas da tabela serão os diferentes tipos de produtos vendidos, como por 
# exemplo: Food and beverages, Fashion accessories, Health and beauty, etc.

# ✅ values='Total' É o campo numérico que será resumido na tabela — nesse caso, o valor total da venda.

# ✅ aggfunc='sum' Função de agregação, que aqui está somando (sum) os valores da coluna Total para cada combinação 
# de Gender e Product line. Poderia ter colocado duas ao mesmo tempo:   aggfunc= ['sum', 'mean']).round(0)

# ✅ .round(0) Após a tabela dinâmica ser criada, o .round(0) arredonda os números para 0 casas decimais, ou seja,
#  sem vírgula, deixando tudo mais limpo.


print (total_gasto )

# Vamos gera o excel sendo que precisamos por nome na planilha(sheet)
total_gasto.to_excel("Total_gasto.xlsx", sheet_name='Total_gasto')

# Desse modo, para automatizar o relatório, precisamos identificar as colunas e linhas mínimas e máximas 
# que ficarão ativas para garantir que, depois de adicionar mais dados à planilha, o código continue 
# funcionando.Carregamos, então, a pasta de trabalho utilizando o load_workbook() e localizamos a planilha
# que queremos trabalhar por meio do wb[‘name_of_sheet’],  por fim, acessamos as células ativas com .active.

wb = load_workbook('Total_gasto.xlsx')
print(wb.sheetnames)
sheet = wb['Total_gasto']


# Referência para a planilha original
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row


