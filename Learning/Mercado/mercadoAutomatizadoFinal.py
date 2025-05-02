# Exercício da página: 
# https://community.revelo.com.br/como-utilizar-python-para-automatizar-planilhas-do-excel/

# Automatizando o relatório em Excel utilizando Python Agora que temos um relatório apresentado por meio de uma
# Tabela Dinâmica, a próxima e mais importante parte é automatização da sua criação. Então, na próxima vez que 
# você quiser fazer esse relatório, basta digitar o nome do arquivo e executá-lo com o código Python.

import pandas as pd
from openpyxl import load_workbook

def automate_excel():
    # Caminho completo do arquivo Excel
    file_name = r'C:\Users\joaorocha\Desktop\Py\Learning\Mercado\supermarket_sales.xlsx'

    # Ler o arquivo Excel
    try:
        
       excel_file = pd.read_excel(file_name)
       print("Arquivo gerado com sucesso")
    except Exception as e:
        print(f"Erro ao carregar o arquivo {e}")
        return

    # Criar a Tabela Dinâmica (Pivot Table)
    
    try: 
        
        report_table = excel_file.pivot_table(
            index='Gender',
            columns='Product line',
            values='Total',
            aggfunc='sum'
        ).round(0)
        print("Tabela dinâmica carregada")
    except Exception as e: 
        print(f"Erro ao criar tabela dinâmica {e}")
        return

  #Exibir uma amostra da tabela dinâmica para verificar se está correta
    print(report_table.head())

    # Extrai a parte "sales.xlsx", ou seja, o nome do mês (sales) e a extensão do arquivo.
    month_and_extension = file_name.split('_')[1]

    # Nome do arquivo para exportar e abrir (evita repetição de código)
    report_file_name = f'report_{month_and_extension}'

    # Exportar a tabela para um novo arquivo Excel
    try:
        report_table.to_excel(
        report_file_name,  # Usando a variável aqui
        sheet_name='Report',  # nomeia a planilha
        startrow = 4  # Começa a escrever a tabela na linha 5 do Excel
    )
        print(f"Arquivo {report_file_name} salvo")  
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel {e}")
        return

    # Abrir o novo arquivo para edição com openpyxl
    try:
        wb = load_workbook(report_file_name)  # Usando a variável aqui
        sheet = wb['Report']
        print(f"Arquivo aberto com sucesso")
    except Exception as e:
        print (f"Erro ao abrir aqrquivo {e}")
        return

    # Pegar as dimensões da planilha ativa
    min_column = sheet.min_column
    max_column = sheet.max_column
    min_row = sheet.min_row
    max_row = sheet.max_row

    # Extrair o nome do mês (sem a extensão)
    month_name = month_and_extension.split('.')[0]

    # Salvar as alterações no arquivo Excel
    try:
        wb.save(report_file_name) 
        print(f"Alterações salvas com sucesso no arquivo: {report_file_name}") 
    except Exception as e:
        print(f"Erro ao salvar as alterações no arquivo Excel {e}")
        return

    # Chamar a função para rodar o processo
automate_excel()

