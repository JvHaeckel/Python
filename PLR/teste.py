from openpyxl import load_workbook
import pandas as pd

# Caminho para o arquivo Excel
arquivo = (r"C:\Users\joaorocha\Desktop\Py\PLR\Projeto Avos - Completo.xlsm")

# Ler o arquivo
# tabela = pd.read_excel(arquivo)

# Nome das abas
aba_base = "Base"
aba_geral = "Geral"

# Carrega o workbook e lê as datas das células Q1 e R1 da aba Base
wb = load_workbook(arquivo, data_only=True)
ws_base = wb[aba_base]
data_inicio = pd.to_datetime(ws_base["Q1"].value)
data_fim = pd.to_datetime(ws_base["R1"].value)

# Lê os dados das abas em DataFrames
base = pd.read_excel(arquivo, sheet_name=aba_base)
geral = pd.read_excel(arquivo, sheet_name=aba_geral)

# Dicionário para armazenar avôs por matrícula
dict_avos = {}

# Exemplo de leitura de dados e cálculo de avôs trabalhados
# Lógica de cálculo dos avôs trabalhados
for i, row in base.iterrows():
    chapa = str(row['Chapa']).strip()
    data_admissao = row['Admissão']
    afastamento = row['Afastamento']
    retorno = row['Retorno']

    avos_trabalhados = 0

    # Loop mês a mês
    for d in pd.date_range(data_inicio, data_fim, freq='MS'):
        primeiro_dia_mes = d.replace(day=1)
        ultimo_dia_mes = (primeiro_dia_mes + pd.offsets.MonthEnd(0))

        if data_admissao > ultimo_dia_mes:
            continue

        ini_trabalho = max(data_admissao, primeiro_dia_mes)
        fim_trabalho = min(ultimo_dia_mes, data_fim)

        # Ajustar se houver afastamento ou retorno
        if pd.notna(afastamento) and afastamento <= fim_trabalho:
            fim_trabalho = min(fim_trabalho, afastamento - pd.Timedelta(days=1))

        if pd.notna(retorno) and retorno >= ini_trabalho and retorno <= ultimo_dia_mes:
            fim_trabalho = max(fim_trabalho, retorno)

        # Calcular dias trabalhados
        if fim_trabalho >= ini_trabalho:
            dias_trabalhados = (fim_trabalho - ini_trabalho).days + 1
            if dias_trabalhados >= 15:
                avos_trabalhados += 1

    # Salvar o cálculo de avôs na planilha Base
    base.at[i, 'Avos Trabalhados'] = avos_trabalhados
    dict_avos[chapa] = avos_trabalhados

# Preencher os dados na aba Geral
for i, row in geral.iterrows():
    chapa_geral = str(row['Chapa']).strip()

    if chapa_geral in dict_avos:
        geral.at[i, 'Avos'] = dict_avos[chapa_geral]
    else:
        # Calcular avôs se não existir na base
        data_admissao_geral = row['Admissão']
        situacao_geral = row['Situação']

        if situacao_geral == 'F':
            geral.at[i, 'Avos'] = 0
        else:
            avos_geral_calculados = 0

            for d in pd.date_range(data_inicio, data_fim, freq='MS'):
                primeiro_dia_mes = d.replace(day=1)
                ultimo_dia_mes = (primeiro_dia_mes + pd.offsets.MonthEnd(0))

                if data_admissao_geral > ultimo_dia_mes:
                    continue

                ini_trabalho = max(data_admissao_geral, primeiro_dia_mes)
                fim_trabalho = min(ultimo_dia_mes, data_fim)

                dias_trabalhados = (fim_trabalho - ini_trabalho).days + 1
                if dias_trabalhados >= 15:
                    avos_geral_calculados += 1

            geral.at[i, 'Avos'] = avos_geral_calculados

# Salvar as alterações de volta nas planilhas
with pd.ExcelWriter(arquivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    writer.book = load_workbook(arquivo, keep_vba=True)
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    base.to_excel(writer, sheet_name=aba_base, index=False)
    geral.to_excel(writer, sheet_name=aba_geral, index=False)
print('Cálculo de avôs finalizado com sucesso!')